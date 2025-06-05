# utils.py 各种工具函数
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side


# 生成3个5-60之间的随机数
def generate_random_number() -> list[int]:
    return [random.randint(5, 60) for _ in range(3)]


def generate_settlement_excel(
    filename: str,
    received_amount: float,
    receivable_items: list,
    company_name: str = "C公司名称"
):
    """
    生成结算单 Excel 文件

    参数：
        filename: 输出的 Excel 文件名，如 "结算单.xlsx"
        received_amount: 实收款项金额（float）
        receivable_items: 应收项目列表，格式如：[("三方/四方货款", 1000), ("第三方费用", 200)]
        company_name: 底部公司名称
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "结算单"

    # 样式
    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def apply_border(cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20

    # 标题
    ws.merge_cells("A1:C1")
    ws["A1"] = "结算单"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # 抬头
    ws.merge_cells("A2:C2")
    ws["A2"] = "（抬头：D公司）"

    # 实收款项标题
    ws.merge_cells("A4:C4")
    ws["A4"] = "实收款项："
    ws["A4"].font = bold_font

    # 实收表头
    ws["A5"] = "序号"
    ws["B5"] = "收款金额（RMB）"
    ws["A5"].font = ws["B5"].font = bold_font
    ws["A5"].alignment = ws["B5"].alignment = center
    apply_border("A5:B5")

    # 实收数据
    ws["A6"] = "1"
    ws["B6"] = received_amount
    ws["B6"].number_format = "#,##0.00"
    apply_border("A6:B6")

    # 应收账款标题
    ws.merge_cells("A8:C8")
    ws["A8"] = "应收账款："
    ws["A8"].font = bold_font

    # 应收表头
    ws["A9"] = "序号"
    ws["B9"] = "项目"
    ws["C9"] = "金额（RMB）"
    for col in ["A", "B", "C"]:
        ws[f"{col}9"].font = bold_font
        ws[f"{col}9"].alignment = center
    apply_border("A9:C9")

    # 应收项目
    for i, (item, amount) in enumerate(receivable_items, start=1):
        row = 9 + i
        ws[f"A{row}"] = i
        ws[f"B{row}"] = item
        ws[f"C{row}"] = amount
        ws[f"C{row}"].number_format = "#,##0.00"
        apply_border(f"A{row}:C{row}")

    subtotal_row = 9 + len(receivable_items) + 1
    ws.merge_cells(f"A{subtotal_row}:B{subtotal_row}")
    ws[f"A{subtotal_row}"] = "小计："
    ws[f"A{subtotal_row}"].alignment = Alignment(horizontal="right")
    ws[f"C{subtotal_row}"] = f"=SUM(C10:C{subtotal_row - 1})"
    ws[f"C{subtotal_row}"].number_format = "#,##0.00"
    apply_border(f"A{subtotal_row}:C{subtotal_row}")

    # 结算款行
    balance_row = subtotal_row + 2
    ws.merge_cells(f"A{balance_row}:B{balance_row}")
    ws[f"A{balance_row}"] = "结算款："
    ws[f"A{balance_row}"].alignment = Alignment(horizontal="right")
    ws[f"C{balance_row}"] = f"=B6 - C{subtotal_row}"
    apply_border(f"A{balance_row}:C{balance_row}")

    # 公司名称
    ws[f"C{balance_row + 2}"] = company_name
    ws[f"C{balance_row + 2}"].alignment = Alignment(horizontal="right")

    # 设置每行高度
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 25

    wb.save(filename)