from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "结算单"

# 样式
title_font = Font(size=14, bold=True)
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def apply_border(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

# 设置列宽
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20

# 第一行：标题
ws.merge_cells("A1:C1")
ws["A1"] = "结算单"
ws["A1"].font = title_font
ws["A1"].alignment = center

# 第二行：抬头
ws.merge_cells("A2:C2")
ws["A2"] = "（抬头：D公司）"
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

# 第三行空一行
# 实收款项标题
ws.merge_cells("A4:C4")
ws["A4"] = "实收款项："
ws["A4"].font = bold_font
apply_border(ws, "A4:C4")

# 表头
ws["A5"] = "序号"
ws["B5"] = "收款金额（RMB）"
ws["A5"].font = ws["B5"].font = bold_font
ws["A5"].alignment = ws["B5"].alignment = center
apply_border(ws, "A5:C5")

# 数据行（留空）
ws["A6"] = "1"
apply_border(ws, "A6:C6")

# 留一空行
# 应收账款标题
ws.merge_cells("A8:C8")
ws["A8"] = "应收账款："
ws["A8"].font = bold_font

# 应收账款表头
ws["A9"] = "序号"
ws["B9"] = "项目"
ws["C9"] = "金额（RMB）"
for col in ["A", "B", "C"]:
    ws[f"{col}9"].font = bold_font
    ws[f"{col}9"].alignment = center
apply_border(ws, "A9:C9")

# 明细项目
items = ["三方/四方货款", "第三方费用", "费用结算服务费", "中标服务费", "购买标书费", "投标服务费"]
for i, item in enumerate(items, start=1):
    row = 9 + i
    ws[f"A{row}"] = i
    ws[f"B{row}"] = item
    apply_border(ws, f"A{row}:C{row}")

# 小计行
subtotal_row = 16
ws.merge_cells(f"A{subtotal_row}:B{subtotal_row}")
ws[f"A{subtotal_row}"] = "小计："
ws[f"A{subtotal_row}"].alignment = Alignment(horizontal="right")
ws[f"C{subtotal_row}"].number_format = "#,##0.00"
ws[f"C{subtotal_row}"].alignment = center
apply_border(ws, f"A{subtotal_row}:C{subtotal_row}")

# 结算款行
# ws.merge_cells("A18:B18")
ws["A18"] = "结算款："
ws["A18"].alignment = Alignment(horizontal="right")
ws["B18"] = "实收款项 - 应收账款 ="
apply_border(ws, "A18:C18")

# 尾注
ws.merge_cells("C20:C20")
ws["C20"] = "C公司名称"
ws["C20"].alignment = Alignment(horizontal="right")

# 设置行高
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 25

# 保存
wb.save("结算单.xlsx")