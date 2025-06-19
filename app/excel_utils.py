import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

import paramiko

import os 
from dotenv import load_dotenv

load_dotenv()

sftp_host = os.getenv("SFTP_HOST")       # 如 szsyjz.vicp.io
sftp_port = os.getenv("SFTP_PORT")                      # 如果你通过 FRP 映射的是其他端口，比如 10022，就写对应端口
sftp_user = os.getenv("SFTP_USER")       # 注意是 SSH 用户，通常是 admin 或你创建的用户
sftp_pass = os.getenv("SFTP_PASS")
# local_file = r"E:\code_projects\syjz_emails\backend\app\scripts\test.txt"        # 本地要上传的文件
remote_path = "JZ/中港模式结算单/"  # 群晖上目标路径，注意要有写权限


def generate_common_settlement_excel(
    filename: str,
    stage: str,
    project_type: str,
    received_amount: float,
    receivable_items: list,
    head_company_name: str,
    bottom_company_name: str
):
    """
    生成结算单 Excel 文件

    参数：
        filename: 输出的 Excel 文件名，如 "结算单.xlsx"
        received_amount: 实收款项金额（float）
        receivable_items: 应收项目列表，格式如：[("三方/四方货款", 1000), ("第三方费用", 200)]
        company_name: 底部公司名称
    """

    # 使用用户主目录作为基础路径
    home_dir = str(Path.home())
    save_dir = os.path.join(home_dir, "settlements")
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "结算单"

    # 样式
    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def apply_border(cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20

    # 标题
    ws.merge_cells("A1:C1")
    ws["A1"] = "结算单"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # 抬头
    ws.merge_cells("A2:C2")
    ws["A2"] = head_company_name

    # 实收款项标题
    ws.merge_cells("A4:C4")
    ws["A4"] = "实收款项："
    ws["A4"].font = bold_font

    apply_border("A4:C4")
    # apply_border("A5:C5")

    # 实收表头
    ws["A5"] = "序号"
    if project_type == "BCD" and stage == "C7":
        ws["B5"] = "中标金额（RMB）"
    else:
        ws["B5"] = "收款金额（RMB）"
    ws["A5"].font = ws["B5"].font = bold_font
    ws["A5"].alignment = ws["B5"].alignment = center
    apply_border("A5:C5")

    # 实收数据
    ws["A6"] = "1"
    ws["B6"] = received_amount
    ws["B6"].number_format = "#,##0.00"
    apply_border("A6:C6")

    # 应收账款标题
    ws.merge_cells("A8:C8")
    ws["A8"] = "应收账款："
    ws["A8"].font = bold_font

    # 应收表头
    ws["A9"] = "序号"
    ws["B9"] = "项目"
    ws["C9"] = "金额（RMB）"
    for col in ["A", "B", "C"]:
        ws[f"{col}9"].font = bold_font
        ws[f"{col}9"].alignment = center
    apply_border("A9:C9")

    # 应收项目
    for i, (item, amount) in enumerate(receivable_items, start=1):
        row = 9 + i
        ws[f"A{row}"] = i
        ws[f"B{row}"] = item
        ws[f"C{row}"] = amount
        ws[f"C{row}"].number_format = "#,##0.00"
        apply_border(f"A{row}:C{row}")

    # 小计行
    subtotal_row = 9 + len(receivable_items) + 1
    receivable_start_row = 10
    receivable_end_row = 9 + len(receivable_items)

    ws.merge_cells(f"A{subtotal_row}:B{subtotal_row}")
    ws[f"A{subtotal_row}"] = "小计："
    ws[f"A{subtotal_row}"].alignment = Alignment(horizontal="right")

    # 修复：确保SUM范围准确
    if receivable_items:
        ws[f"C{subtotal_row}"] = f"=SUM(C{receivable_start_row}:C{receivable_end_row})"
    else:
        ws[f"C{subtotal_row}"] = 0

    ws[f"C{subtotal_row}"].number_format = "#,##0.00"
    apply_border(f"A{subtotal_row}:C{subtotal_row}")

    # 结算款行
    balance_row = subtotal_row + 2
    ws.merge_cells(f"A{balance_row}:B{balance_row}")
    ws[f"A{balance_row}"] = "结算款(RMB)："
    ws[f"A{balance_row}"].alignment = Alignment(horizontal="right")
    ws[f"C{balance_row}"] = f"=B6 - C{subtotal_row}"  # 正确的结算公式
    ws[f"C{balance_row}"].number_format = "#,##0.00"
    apply_border(f"A{balance_row}:C{balance_row}")

    # 公司名称
    ws[f"C{balance_row + 2}"] = bottom_company_name
    ws[f"C{balance_row + 2}"].alignment = Alignment(horizontal="right")

    # 设置每行高度
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    wb.save(file_path)
    
    #TODO 上传文件到共享服务器
    try:
        transport = paramiko.Transport((sftp_host, sftp_port))
        transport.connect(username=sftp_user, password=sftp_pass)
        print("✅ FTP连接成功")
        sftp = paramiko.SFTPClient.from_transport(transport)
        remote_pathfile = f"{remote_path}/{filename}"
        print("local file: ", file_path)
        print("remote path: ", remote_pathfile)

        sftp.put(file_path, remote_pathfile)

        print("✅ 文件上传成功")
        sftp.close()
        transport.close()

    except Exception as e:
        print("❌ 上传失败:", str(e))



    return file_path